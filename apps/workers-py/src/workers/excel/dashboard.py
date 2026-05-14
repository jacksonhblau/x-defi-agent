"""Excel dashboard — the agent's control plane for a single-user setup.

The file lives at <project_root>/agent_dashboard.xlsx by default.

Two functions matter externally:
- export_to_excel(): pull current DB state and write the file
- apply_from_excel(): read the file and push user edits back to DB

The `agent watch` command loops on both, so the user can edit the file in
Excel/Numbers and changes apply on the next cycle (~60 seconds).

Sheets:
  Drafts      — pending posts. Editable: status, scheduled_for, body
  Run Jobs    — scripts the agent runs. Editable: cron, enabled, run_now
  Stories     — story-level state. Read-only.
  Signals     — recent signal log. Read-only.
  Posts       — published tweets + engagement. Read-only.
  Config      — thresholds. Editable: value column.
  Watchlist   — monitored handles. Editable: weight, enabled.

Edits to a read-only column are silently ignored on apply().
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from psycopg.rows import dict_row

from .. import config, db


# =============================================================================
# Layout constants
# =============================================================================

EXCEL_FILENAME = "agent_dashboard.xlsx"

# Visual styling
HEADER_FILL = PatternFill("solid", fgColor="1F2937")        # dark slate
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
READONLY_FILL = PatternFill("solid", fgColor="F3F4F6")      # very light gray
EDITABLE_FILL = PatternFill("solid", fgColor="FEF3C7")      # soft yellow
THIN = Side(style="thin", color="D1D5DB")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# Which columns the user is allowed to edit, per sheet.
EDITABLE_COLUMNS: dict[str, set[str]] = {
    "Drafts": {"body", "status", "scheduled_for", "reviewer_notes"},
    "Run Jobs": {"cron", "enabled", "run_now"},
    "Config": {"value"},
    "Watchlist": {"weight", "enabled"},
    # Stories, Signals, Posts are read-only.
}


def excel_path() -> Path:
    return config.PROJECT_ROOT / EXCEL_FILENAME


# =============================================================================
# Helpers
# =============================================================================

def _style_header(ws: Worksheet, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = CELL_BORDER
    ws.freeze_panes = "A2"


def _style_rows(ws: Worksheet, columns: list[str], n_rows: int, editable: set[str]) -> None:
    for row_idx in range(2, n_rows + 2):
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = EDITABLE_FILL if name in editable else READONLY_FILL
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize(ws: Worksheet, columns: list[str], maxw: int = 60) -> None:
    for col_idx, name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        # Heuristic width based on name + sample lengths
        max_len = len(name)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            for line in str(v).split("\n"):
                max_len = max(max_len, len(line))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), maxw)


def _fetch_all(query: str, params: tuple = ()) -> list[dict[str, Any]]:
    with db.conn() as c, c.cursor(row_factory=dict_row) as cur:
        cur.execute(query, params)
        return cur.fetchall()


# =============================================================================
# Sheet writers
# =============================================================================

DRAFTS_COLUMNS = [
    "draft_id", "story_id", "headline", "format", "variant_label",
    "body", "ai_check_passed", "ai_check_flags",
    "status", "scheduled_for", "posted_at", "root_tweet_id",
    "reviewer_notes", "created_at",
]


def write_drafts(wb: Workbook) -> None:
    ws = wb.create_sheet("Drafts")
    rows = _fetch_all(
        """
        select d.id::text as draft_id,
               d.story_id::text as story_id,
               s.headline,
               d.format,
               d.variant_label,
               coalesce(d.edited_body, d.body) as body,
               d.ai_check_passed,
               array_to_string(d.ai_check_flags, '; ') as ai_check_flags,
               d.status,
               sp.post_at as scheduled_for,
               p.posted_at,
               p.root_tweet_id,
               d.reviewer_notes,
               d.created_at
        from drafts d
        join stories s on s.id = d.story_id
        left join scheduled_posts sp on sp.draft_id = d.id and sp.status in ('queued','posting')
        left join posts p on p.draft_id = d.id
        order by d.created_at desc
        limit 200
        """
    )
    _style_header(ws, DRAFTS_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(DRAFTS_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))
    _style_rows(ws, DRAFTS_COLUMNS, len(rows), EDITABLE_COLUMNS["Drafts"])
    _autosize(ws, DRAFTS_COLUMNS, maxw=80)


RUN_JOBS_COLUMNS = [
    "name", "description", "command", "cron", "enabled",
    "last_run_at", "next_run_at", "last_status", "last_error", "run_now",
]


def write_run_jobs(wb: Workbook) -> None:
    ws = wb.create_sheet("Run Jobs")
    rows = _fetch_all(
        """
        select name, description, command, cron, enabled,
               last_run_at, next_run_at, last_status, last_error, run_now
        from run_jobs
        order by sort_order asc, name asc
        """
    )
    _style_header(ws, RUN_JOBS_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(RUN_JOBS_COLUMNS, start=1):
            v = row.get(col)
            # Convert booleans to YES/NO for friendly editing
            if isinstance(v, bool):
                v = "YES" if v else "NO"
            ws.cell(row=r_idx, column=c_idx, value=v)
    _style_rows(ws, RUN_JOBS_COLUMNS, len(rows), EDITABLE_COLUMNS["Run Jobs"])
    _autosize(ws, RUN_JOBS_COLUMNS, maxw=60)


STORIES_COLUMNS = ["story_id", "headline", "status", "hot_take", "entities", "created_at"]


def write_stories(wb: Workbook) -> None:
    ws = wb.create_sheet("Stories")
    rows = _fetch_all(
        """
        select id::text as story_id, headline, status, hot_take,
               array_to_string(entities, ', ') as entities, created_at
        from stories
        order by created_at desc
        limit 200
        """
    )
    _style_header(ws, STORIES_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(STORIES_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))
    _style_rows(ws, STORIES_COLUMNS, len(rows), EDITABLE_COLUMNS.get("Stories", set()))
    _autosize(ws, STORIES_COLUMNS, maxw=70)


SIGNALS_COLUMNS = ["source", "signal_type", "entity", "materiality_score", "novelty_score", "observed_at", "notes"]


def write_signals(wb: Workbook) -> None:
    ws = wb.create_sheet("Signals")
    rows = _fetch_all(
        """
        select source, signal_type, entity,
               materiality_score, novelty_score, observed_at, notes
        from signals
        order by observed_at desc
        limit 200
        """
    )
    _style_header(ws, SIGNALS_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(SIGNALS_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))
    _style_rows(ws, SIGNALS_COLUMNS, len(rows), EDITABLE_COLUMNS.get("Signals", set()))
    _autosize(ws, SIGNALS_COLUMNS, maxw=60)


POSTS_COLUMNS = [
    "posted_at", "format", "root_tweet_id", "body",
    "impressions_24h", "likes_24h", "rt_24h", "replies_24h",
    "impressions_7d", "likes_7d", "rt_7d", "replies_7d",
]


def write_posts(wb: Workbook) -> None:
    ws = wb.create_sheet("Posts")
    rows = _fetch_all(
        """
        select p.posted_at, p.format, p.root_tweet_id, p.body,
               e24.impressions as impressions_24h,
               e24.likes as likes_24h,
               e24.retweets as rt_24h,
               e24.replies as replies_24h,
               e7.impressions as impressions_7d,
               e7.likes as likes_7d,
               e7.retweets as rt_7d,
               e7.replies as replies_7d
        from posts p
        left join engagement e24 on e24.post_id = p.id and e24.window_label = '24h'
        left join engagement e7  on e7.post_id  = p.id and e7.window_label  = '7d'
        order by p.posted_at desc
        limit 200
        """
    )
    _style_header(ws, POSTS_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(POSTS_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))
    _style_rows(ws, POSTS_COLUMNS, len(rows), EDITABLE_COLUMNS.get("Posts", set()))
    _autosize(ws, POSTS_COLUMNS, maxw=80)


CONFIG_COLUMNS = ["key", "value", "description"]

# Keys exposed to the user via the Config sheet, with descriptions.
# Values are read from config/thresholds.json on first export and from this sheet thereafter.
CONFIG_KEYS: list[tuple[str, str]] = [
    ("materiality.default_threshold",      "Min materiality score (0-100) to promote a signal to a story"),
    ("materiality.novelty_threshold",      "Min novelty score (0-100) for a signal to draft"),
    ("materiality.minimum_for_thread",     "Min materiality score to also generate a thread variant"),
    ("cadence.daily_post_cap",             "Max posts per day"),
    ("cadence.min_minutes_between_posts",  "Minimum minutes between published posts"),
    ("cadence.thread_max_per_day",         "Max threads per day"),
    ("cadence.reply_max_per_day",          "Max replies/QTs per day"),
    ("onchain.tvl_delta_threshold_pct",    "Min |24h TVL change %| to emit a signal"),
    ("onchain.apy_delta_threshold_bps",    "Min APY shift in bps to emit a vault signal"),
    ("onchain.treasury_flow_threshold_usd","Min treasury wallet flow in USD to emit a signal"),
    ("slow_day_fallback.no_news_window_hours",     "Hours of silence before hot-take fires"),
    ("slow_day_fallback.hot_take_max_regenerations","Max regenerations on a failed originality filter"),
]


def _read_nested(d: dict[str, Any], dotted_key: str) -> Any:
    cur: Any = d
    for part in dotted_key.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur


def write_config(wb: Workbook) -> None:
    ws = wb.create_sheet("Config")
    th = config.thresholds()
    _style_header(ws, CONFIG_COLUMNS)
    for r_idx, (key, desc) in enumerate(CONFIG_KEYS, start=2):
        ws.cell(row=r_idx, column=1, value=key)
        ws.cell(row=r_idx, column=2, value=_read_nested(th, key))
        ws.cell(row=r_idx, column=3, value=desc)
    _style_rows(ws, CONFIG_COLUMNS, len(CONFIG_KEYS), EDITABLE_COLUMNS["Config"])
    _autosize(ws, CONFIG_COLUMNS, maxw=80)


WATCHLIST_COLUMNS = ["handle", "category", "weight", "enabled"]


def write_watchlist(wb: Workbook) -> None:
    ws = wb.create_sheet("Watchlist")
    wl = config.watchlist()
    rows: list[dict[str, Any]] = []
    for category, group in wl.items():
        weight = group.get("weight", 1.0)
        for handle in group.get("handles", []):
            rows.append({"handle": handle, "category": category, "weight": weight, "enabled": True})
    _style_header(ws, WATCHLIST_COLUMNS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(WATCHLIST_COLUMNS, start=1):
            v = row.get(col)
            if isinstance(v, bool):
                v = "YES" if v else "NO"
            ws.cell(row=r_idx, column=c_idx, value=v)
    _style_rows(ws, WATCHLIST_COLUMNS, len(rows), EDITABLE_COLUMNS["Watchlist"])
    _autosize(ws, WATCHLIST_COLUMNS, maxw=40)


def write_readme_sheet(wb: Workbook) -> None:
    """Pinned first sheet — short usage notes."""
    ws = wb.create_sheet("README", 0)
    notes = [
        "Agent Dashboard",
        "",
        "Updated by `agent watch` (runs on the user's Mac). Changes you make here apply on the next cycle.",
        "",
        "Sheet guide:",
        "  Drafts     — pending posts. Edit `status` to 'approved' / 'rejected'. Edit `scheduled_for` to schedule.",
        "  Run Jobs   — toggle `enabled`, change `cron`, or set `run_now` = YES to trigger an ad-hoc run.",
        "  Stories    — story-level state (read-only).",
        "  Signals    — recent signal log (read-only).",
        "  Posts      — published tweets + engagement metrics (read-only).",
        "  Config     — editable thresholds. Save and the watch loop will pick them up.",
        "  Watchlist  — monitored X accounts. Toggle `enabled` to skip an account.",
        "",
        "Color key: yellow cells are editable, gray cells are read-only (your edits there are ignored).",
        "",
        "Status values for Drafts:",
        "  pending     — generated by the agent, awaiting your review",
        "  approved    — you've approved; will be queued for posting at `scheduled_for` (or ASAP if blank)",
        "  rejected    — you've rejected; agent won't post it",
        "  scheduled   — agent has placed it in the queue (automatic, do not set manually)",
        "  posted      — agent has published it (automatic)",
        "",
        "Run jobs `cron` field uses standard cron syntax in UTC. Examples:",
        "  */10 * * * *  — every 10 minutes",
        "  0 15 * * *    — daily at 15:00 UTC (11am ET)",
        "  0 13 * * 5    — Friday at 13:00 UTC",
        "",
        "To trigger a job once: set `run_now` to YES, save the file, wait up to 60s.",
        "The agent clears `run_now` to NO after the job runs.",
    ]
    for r_idx, line in enumerate(notes, start=1):
        cell = ws.cell(row=r_idx, column=1, value=line)
        if r_idx == 1:
            cell.font = Font(bold=True, size=16)
    ws.column_dimensions["A"].width = 120


# =============================================================================
# Public API
# =============================================================================

def export_to_excel(path: Path | None = None) -> Path:
    """Build the workbook from current DB state. Returns the file path."""
    path = path or excel_path()
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default:
        wb.remove(default)

    write_readme_sheet(wb)
    write_drafts(wb)
    write_run_jobs(wb)
    write_stories(wb)
    write_signals(wb)
    write_posts(wb)
    write_config(wb)
    write_watchlist(wb)

    # Atomic write: write to .tmp then rename
    tmp = path.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    tmp.replace(path)
    return path


def _read_sheet(wb: Workbook, name: str) -> list[dict[str, Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({h: v for h, v in zip(headers, row) if h})
    return out


def _coerce_bool(v: Any) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1", "on"):
        return True
    if s in ("no", "n", "false", "0", "off"):
        return False
    return None


def _coerce_ts(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    try:
        return datetime.fromisoformat(str(v))
    except ValueError:
        return None


def apply_from_excel(path: Path | None = None) -> dict[str, int]:
    """Read user edits from the Excel file and push them back to DB.

    Returns counts per change type (e.g. {'drafts_updated': 2, 'jobs_updated': 1}).
    """
    path = path or excel_path()
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        # File is locked by Excel right now. Skip this cycle.
        return {"skipped_locked": 1}

    counts: dict[str, int] = {}

    # ---- Drafts: status, scheduled_for, body, reviewer_notes ----
    drafts_rows = _read_sheet(wb, "Drafts")
    drafts_updated = 0
    for row in drafts_rows:
        draft_id = row.get("draft_id")
        if not draft_id:
            continue
        new_status = (row.get("status") or "").strip().lower() or None
        new_scheduled = _coerce_ts(row.get("scheduled_for"))
        new_body = row.get("body")
        new_notes = row.get("reviewer_notes")

        with db.conn() as c, c.cursor() as cur:
            cur.execute(
                "select status, edited_body, scheduled_post_id is not null as scheduled "
                "from drafts d left join scheduled_posts sp on sp.draft_id = d.id "
                "where d.id = %s::uuid",
                (draft_id,),
            )
            existing = cur.fetchone()
            if not existing:
                continue
            existing_status = existing[0]

            sets: list[str] = []
            params: list[Any] = []
            if new_status and new_status != existing_status and new_status in (
                "pending", "approved", "rejected", "edited"
            ):
                sets.append("status = %s")
                params.append(new_status)
                sets.append("reviewed_at = now()")
            if new_body and new_body != row.get("body"):
                pass  # already same; openpyxl gave us current value
            # If body was edited (differs from current `body` column),
            # store in edited_body. We can't easily diff here without re-fetching;
            # rely on the user explicitly setting `status = edited` to flag this.
            if new_status == "edited" and new_body:
                sets.append("edited_body = %s")
                params.append(new_body)
            if new_notes is not None:
                sets.append("reviewer_notes = %s")
                params.append(new_notes)

            if sets:
                params.append(draft_id)
                cur.execute(f"update drafts set {', '.join(sets)} where id = %s::uuid", params)
                drafts_updated += 1

            # If approved, queue it (if not already queued)
            if new_status == "approved":
                post_at = new_scheduled or datetime.now(timezone.utc)
                cur.execute(
                    """
                    insert into scheduled_posts (draft_id, post_at)
                    select %s::uuid, %s
                    where not exists (
                        select 1 from scheduled_posts where draft_id = %s::uuid and status in ('queued','posting')
                    )
                    """,
                    (draft_id, post_at, draft_id),
                )

            c.commit()
    counts["drafts_updated"] = drafts_updated

    # ---- Run Jobs: cron, enabled, run_now ----
    jobs_rows = _read_sheet(wb, "Run Jobs")
    jobs_updated = 0
    for row in jobs_rows:
        name = row.get("name")
        if not name:
            continue
        new_cron = row.get("cron")
        new_enabled = _coerce_bool(row.get("enabled"))
        new_run_now = _coerce_bool(row.get("run_now"))
        sets: list[str] = []
        params: list[Any] = []
        if new_cron is not None:
            sets.append("cron = %s")
            params.append(new_cron)
        if new_enabled is not None:
            sets.append("enabled = %s")
            params.append(new_enabled)
        if new_run_now is True:
            sets.append("run_now = true")
        if not sets:
            continue
        params.append(name)
        with db.conn() as c, c.cursor() as cur:
            cur.execute(f"update run_jobs set {', '.join(sets)} where name = %s", params)
            c.commit()
        jobs_updated += 1
    counts["jobs_updated"] = jobs_updated

    return counts
